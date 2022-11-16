import openpyxl
book = openpyxl.load_workbook('D:/Sushmitha/framework/pythonProject/TestData/Book1.xlsx')
sheet = book.active
cell = sheet.cell(row=1, column=2)   #another way print(sheet['A1'].value)
print(cell.value)

sheet.cell(row=4, column=3).value = "Sweden"
print(sheet.cell(row=4, column=3).value)

print(sheet.max_row)
print(sheet.max_column)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "ge":
        for j in range(1, sheet.max_column+1):
            print(sheet.cell(row=i, column=j).value)
            
